import openpyxl
import math
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import numpy as np
import random
import os
import decimal

# Константы
ro = 15316.0 # км
gr2rad = math.pi/180.0 # Перевод градусоы в радианы
arg = 90 * gr2rad # аргумент для косинуса 90 град. в радианах

# --- Конфигурация таблицы ---
NUM_FILES = 30
NUM_ROWS = NUM_FILES  # OY (строки, от 0 до NUM_FILES)
NUM_COLS = NUM_FILES  # OX NUM_FILES(столбцы, от 0 до NUM_FILES)
# --- Файлы, необходимые для работы ---
DATA_FILE = 'C:/Users/User/selenium_course/data_file.dat' # Имя файла с данными для ячеек в Excel (файл создается)
EXCEL_FILE = 'C:/Users/User/selenium_course/colored_table.xlsx' # Имя выходного Excel файла
FILE_NAME = 'C:/Users/User/selenium_course/efem30ckkvkk/files/name_file.dat' # Имя файла с названиями файлов со значениями для КА
FILE_PATH = 'C:/Users/User/selenium_course/efem30ckkvkk/files/' # Путь к файлам со значениями для КА

# --- Цвета в формате ARGB (Alpha, Red, Green, Blue) для openpyxl ---
# FF - полная непрозрачность
RED_HEX = "FFFF0000"
BLACK_HEX = "FF000000"
WHITE_HEX = "FFFFFFFF"

# ---  Генерируем цвета из градиента от красного к зеленому через желтый. 100 цветов.
def generate_red_to_green_gradient_argb():
    """
    Генерирует список ARGB цветов для градиента от красного к зеленому,
    проходя через желтый, на шкале от 0 до 100.
    """
    colors = []
    max_rgb_value = 255 # Максимальное значение для каждого цветового канала (0-255)

    for i in range(101): # От 0 до 100 включительно
        r, g, b = 0, 0, 0

        if i <= 50:
            # Сегмент от Красного к Желтому
            # Красный остается максимальным
            r = max_rgb_value
            # Зеленый увеличивается от 0 до 255
            # Прогресс в этом сегменте: i / 50
            g = int(i / 50 * max_rgb_value)
        else: # i > 50
            # Сегмент от Желтого к Зеленому
            # Зеленый остается максимальным
            g = max_rgb_value
            # Красный уменьшается от 255 до 0
            # Прогресс в этом сегменте: (i - 50) / 50
            # Мы вычитаем этот прогресс из 255
            r = int(max_rgb_value - ((i - 50) / 50 * max_rgb_value))

        # Убедимся, что значения находятся в диапазоне 0-255
        r = max(0, min(max_rgb_value, r))
        g = max(0, min(max_rgb_value, g))
        b = max(0, min(max_rgb_value, b)) # Синий всегда 0 в этом градиенте

        # Форматируем в ARGB (Alpha: FF - полная непрозрачность)
        # :02X означает форматирование числа как шестнадцатеричное с двумя символами,
        # дополненными нулями при необходимости (например, 0 станет 00, 10 станет 0A)
        argb_hex = f"FF{r:02X}{g:02X}{b:02X}"
        colors.append(argb_hex)

    return colors

# --- Функции для работы с данными ---
def dot_product(v1, v2):
    """Вычисляет скалярное произведение двух векторов."""
    return v1[0]*v2[0] + v1[1]*v2[1] + v1[2]*v2[2]

def module_vector(v):
    """Вычисляет модуль вектора."""
    # Избегаем sqrt(отрицательное число) из-за погрешностей float
    return math.sqrt(max(0.0, (v[0]**2 + v[1]**2 + v[2]**2)))

def subtract_vectors(v1, v2):
    """Вычитает вектор v2 из вектора v1."""
    return [v1[0] - v2[0], v1[1] - v2[1], v1[2] - v2[2]]

def usl_visability(v1, v2):
    """Вычисляет условия видимости через скалярное произведение векторов (v1-v2)*v1."""
    global ro, arg
    """Вычисляет косинусы углов"""
    cos_betta = math.cos(arg) # 180/2 = 90 в радианах
    cos_alpha_ij = (dot_product(subtract_vectors(v1, v2), v1) / module_vector(subtract_vectors(v1, v2))) / module_vector(v1)
    cos_alpha_ji = (dot_product(subtract_vectors(v2, v1), v2) / module_vector(subtract_vectors(v2, v1))) / module_vector(v2)
    """Вычисляет условия видимости"""
    if 1.0 - cos_alpha_ij**2 > 0:
        usl1 = module_vector(v1) * math.sqrt(1.0 - cos_alpha_ij**2) - ro
    else:
        usl1 = -1 # Если cos_alpha_ij ~ 1, то получается корень из отрицательного числа, спутники друг друга не видят, поэтому присваиваем -1
    usl2 = cos_alpha_ij - cos_betta
    usl3 = cos_alpha_ji - cos_betta
    return usl1, usl2, usl3
                                   
def read_satellite_data(namefiles_path=FILE_NAME):
    """
    Считывает названия файлов из namefiles.txt, затем открывает каждый из них
    и считывает вещественные значения с 3 по 6 столбец (время и координаты КА), начиная со второй строки.

    Возвращает словарь, где ключи - это имена файлов, а значения - списки,
    содержащие списки из 4 вещественных чисел (столбцы 3-6) для каждой строки.
    """
    global nstroki
    
    try:
        with open(namefiles_path, 'r') as f:
            filenames = [(FILE_PATH + line).strip() for line in f if line.strip()] # Читаем имена файлов, удаляя пробелы и пустые строки
    except FileNotFoundError:
        print(f"Ошибка: Файл '{namefiles_path}' не найден.")
        return None
    
    print(f"\nСчитываем данные из {len(filenames)} файлов.")

    all_files_data_rows = [] # Список для хранения данных из каждого файла

    for file_idx, filename in enumerate(filenames): # file_idx - индекс файла
        current_file_data = [] # Список для хранения строк [x, y, z] текущего файла

        try:
            with open(filename, 'r') as f:
                # Пропускаем первые две строки заголовка
                try:
                    next(f) 
                    next(f)
                except StopIteration:
                    print(f"Предупреждение: Файл '{filename}' содержит менее двух строк. Возможно, он пуст.")
                    # Продолжаем, current_file_data останется пустым

                # Считываем остальные строки
                # line_num_in_file отсчитывается с 3 для отображения исходного номера строки
                for line_num_in_file, line in enumerate(f, start=3): 
                    parts = line.split() 
                    if len(parts) >= 6: # Убедимся, что есть как минимум 6 столбцов
                        try:
                            # Предполагаем, что x, y, z находятся в столбцах 3, 4, 5 (индексы 3, 4, 5)
                            x = float(parts[3])
                            y = float(parts[4])
                            z = float(parts[5])
                            current_file_data.append([x, y, z])
                        except ValueError:
                            print(f"Предупреждение: Не удалось преобразовать данные в float в файле '{filename}', строка {line_num_in_file}: '{line.strip()}'")
                    else:
                        print(f"Предупреждение: Недостаточно столбцов в файле '{filename}', строка {line_num_in_file}: '{line.strip()}'")

            if current_file_data:
                #print(f"Успешно прочитано {len(current_file_data)} строк данных из '{filename}'")
                all_files_data_rows.append(np.array(current_file_data))
            else:
                print(f"Из файла '{filename}' не удалось прочитать валидные данные. Добавлен пустой массив.")
                all_files_data_rows.append(np.array([]).reshape(0,3)) # Добавляем пустой 2D массив, чтобы сохранить количество файлов

        except FileNotFoundError:
            print(f"Ошибка: Файл данных '{filename}' не найден. Пропускаю и добавляю пустой массив.")
            all_files_data_rows.append(np.array([]).reshape(0,3))
        except Exception as e:
            print(f"Произошла непредвиденная ошибка при чтении '{filename}': {e}. Добавляю пустой массив.")
            all_files_data_rows.append(np.array([]).reshape(0,3))

    if not all_files_data_rows:
        print("После обработки всех файлов данные не были собраны.")
        return None

    # Теперь all_files_data_rows - это список 2D numpy массивов.
    # Чтобы получить один 3D массив, нужно учесть, что файлы могут содержать разное количество строк.
    # Заполняем пропущенные значения NaN.

    # Определяем максимальное количество строк среди всех файлов
    max_rows_per_file = 0
    # Проверяем, что список не пуст перед поиском максимума
    if all_files_data_rows:
        row_counts = [arr.shape[0] for arr in all_files_data_rows if arr.shape[0] > 0]
        if row_counts:
            max_rows_per_file = max(row_counts)

    if max_rows_per_file == 0:
        print("Все прочитанные файлы не содержали валидных данных.")
        return np.array([]).reshape(len(all_files_data_rows), 0, 3) # Возвращаем 3D массив с 0 строк данных

    # Инициализируем итоговый 3D NumPy массив значениями NaN для заполнения
    # Размеры: (количество_файлов, максимальное_количество_строк_в_любом_файле, 3 для x,y,z)
    final_sat_data = np.full(
        (len(all_files_data_rows), max_rows_per_file, 3), 
        np.nan, 
        dtype=float
    )

    # Заполняем итоговый 3D массив данными
    for file_idx, file_arr in enumerate(all_files_data_rows):
        if file_arr.shape[0] > 0: # Если в этом файле есть данные
            final_sat_data[file_idx, :file_arr.shape[0], :] = file_arr

    nstroki = len(all_files_data_rows)
    
    return final_sat_data

def generate_visability_data_file(filepath, rows, cols, numfile):
#    """
#    Вычисляем вероятности радиовидимости КА сохраняет их в файл в виде матрицы.
#    """
    global nstroki
# Cчитываем данные спутников из файлов, записываем в словарь satellite_data внутри функции read_satellite_data()
    satellite_data = read_satellite_data()

    matr_vidim = [[int(0)]*rows for _ in range(cols)] # создаем матрицу с нулевым элементами (cols и rows у нас всегда равны)
    
    if satellite_data.all():
        print("\n--- Проверяем условия видимости ---")
        
        # # Перебираем данные из файлов. Проверяем условия видимости.
        for i in range(numfile):
            for j in range(numfile):
                print('i_file =', i,'j_file =', j)
                count = 0
                if i != j:
                    for k in range(nstroki):
                        usl = usl_visability(satellite_data[i][k][0:3], satellite_data[j][k][0:3])
                        if usl[0] > 0 and usl[1] > 0 and usl[2] > 0:
                            count += 1 # количество видимостей за период времени
                    matr_vidim[i][j] = int(100.0 * count / nstroki + 0.5)
                else:
                    matr_vidim[i][j] = 0
            vidimost = np.array(matr_vidim)
        np.savetxt(filepath, vidimost, fmt='%d', delimiter=' ')
  
def load_data_from_file(filepath, expected_rows, expected_cols):
    """
    Загружает полученные данные видимости из файла.
    Проверяет, соответствуют ли размеры данных ожидаемым.
    """
    if not os.path.exists(filepath):
        print(f"Ошибка: Файл данных '{filepath}' не найден.")
        return None

    try:
        data = np.loadtxt(filepath, dtype=int)
    except ValueError:
        print(f"Ошибка: Не удалось прочитать данные из файла '{filepath}'. Убедитесь, что он содержит только числа.")
        return None
    except Exception as e:
        print(f"Произошла ошибка при чтении файла: {e}")
        return None

    # Проверяем, что данные являются двумерным массивом
    if data.ndim != 2:
        print(f"Ошибка: Данные в файле '{filepath}' не являются двумерной матрицей. Обнаружено измерений: {data.ndim}.")
        return None

    if data.shape != (expected_rows, expected_cols):
        print(f"Ошибка: Размеры данных в файле ({data.shape}) не соответствуют ожидаемым ({expected_rows}, {expected_cols}).")
        print("Пожалуйста, убедитесь, что файл содержит нужное количество строк и столбцов.")
        return None
    return data

# --- Функция для определения цвета ячейки ---

def get_cell_colors(value, is_diagonal):
    """
    Определяет цвет заливки и цвет текста для ячейки на основе ее значения.
    Возвращает кортеж (fill_color_hex, font_color_hex).
    """
    global mapped_data
    
    if is_diagonal:
        return BLACK_HEX, WHITE_HEX # Черный фон, белый текст
    
    if value in mapped_data:
        color = mapped_data[value] # Берем значение из словаря
        #print(color)
        return color, BLACK_HEX # Красный фон, белый текст (для контраста)

# --- Функция для построения и сохранения Excel таблицы ---

def create_excel_table(data, excel_filepath, rows, cols):
    """
    Создает Excel файл с таблицей, раскрашенной по значениям.
    """
    
    print(f"Создание Excel файла: {excel_filepath}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Colored Table"

    # Определяем стиль двойной границы и тонкой границы
    double_line_side = Side(border_style="double", color=BLACK_HEX) # Двойная
    thin_line_side = Side(border_style='thin', color=BLACK_HEX) # Тонкая для остальных сторон
         
    # Определяем ряды и колонки, где должны быть границы
    # Границы будут после 8-й, 16-й и 24-й ячейки
    # И перед 1-й, 9-й, 17-й ячейками
    top_border_rows = {0, 8, 16}
    bottom_border_rows = {7, 15, 23, 29}
    left_border_cols = {0, 8, 16}
    right_border_cols = {7, 15, 23, 29}
        
    # Заголовки столбцов (OX)
    # Начинаем с A1 пустой, затем B1, C1... для 0, 1, 2...
    ws.cell(row=1, column=1).value = "№" # Пустая ячейка A1
    ws.column_dimensions[get_column_letter(1)].width = 5 # Ширина первого столбца 5 символов
    for j in range(cols):
        col_idx = j + 2 # Excel колонки 1-индексированы, и мы смещаемся на 1 для заголовка OY
        ws.cell(row=1, column=col_idx).value = j + 1 # +1 чтобы строки нумеровались, начиная с 1, а не с 0
        # Устанавливаем ширину столбца для лучшей читаемости
        ws.column_dimensions[get_column_letter(col_idx)].width = 5 # Например, 5 символов

    # Заголовки строк (OY)
    # Начинаем с A2, A3... для 0, 1, 2...
    for i in range(rows):
        row_idx = i + 2 # Excel строки 1-индексированы, и мы смещаемся на 1 для заголовка OX
        ws.cell(row=row_idx, column=1).value = i + 1 # +1 чтобы строки нумеровались, начиная с 1, а не с 0
        ws.row_dimensions[row_idx].height = 20 # Устанавливаем высоту строки
    
    # --- Делаем границу у ячеек 1 строки и 1 столбца таблицы
    for i in range(1, rows+2):
        cell = ws.cell(row=i, column=1)
        cell.border = Border(left=thin_line_side,
                             right=thin_line_side,
                             top=thin_line_side,
                             bottom=thin_line_side)
    for j in range(1, cols+2):
        cell = ws.cell(row=1, column=j)    
        cell.border = Border(left=thin_line_side,
                             right=thin_line_side,
                             top=thin_line_side,
                             bottom=thin_line_side)
             
    # --- Заполнение таблицы данными и применение форматирования ---
    for i in range(rows):
        for j in range(cols):
            value = data[i, j]
            # Ячейка данных находится по смещению +2 от индексов массива
            # (1 для заголовка OX, 1 для заголовка OY)
            cell = ws.cell(row=i + 2, column=j + 2)
            cell.value = value

            # Проверяем, является ли ячейка на главной диагонали
            # Главная диагональ только для квадратной части таблицы
            is_diagonal = (i == j)
            
            fill_hex, font_hex = get_cell_colors(value, is_diagonal) # определяет цвет ячейки относительно значения

            cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            cell.font = Font(color=font_hex, bold=True) # Жирный шрифт для лучшей читаемости

            # Инициализируем компоненты границы для текущей ячейки
            # По умолчанию - граница тонкая линия
            top_b = thin_line_side
            bottom_b = thin_line_side
            left_b = thin_line_side
            right_b = thin_line_side
                
            # Определяем, нужна ли двойная граница с каждой стороны ячейки
            if i in top_border_rows:
                top_b = double_line_side
            if i in bottom_border_rows:
                bottom_b = double_line_side
            if j in left_border_cols:
                left_b = double_line_side
            if j in right_border_cols:
                right_b = double_line_side

            # Применяем собранную границу к ячейке
            cell.border = Border(left=left_b,
                                 right=right_b,
                                 top=top_b,
                                 bottom=bottom_b)       

    # Определяем список римских цифр
    rim_num = ['I', 'II', 'III', 'IV']

    # --- Двойные границы для столбцов, под таблицей. Определяем константы, которые не меняются в цикле
    base_start_row = rows + 2  # Начинаем с 32 строки (rows + 2)
    num_rows_merge = 3
    num_cols_merge = 8   # +8 для орбит.группировки из 8 КА
    num_cols_merge_6 = 6 # +6 для орбит.группировки из 6 КА
    end_row_merge = base_start_row + num_rows_merge - 1

    # Создаем объекты стилей один раз, чтобы не пересоздавать их в каждой итерации
    font_style = Font(size=16, bold=True)
    alignment_style = Alignment(horizontal='center', vertical='center')
    border_style = Border(left=double_line_side,
                          right=double_line_side,
                          top=double_line_side,
                          bottom=double_line_side)

    # Начальный столбец для первой объединенной ячейки
    initial_start_col = 2

    # Проходимся по списку римских цифр
    for i, roman_value in enumerate(rim_num):
        # Рассчитываем начальный столбец для текущей итерации
        # start_col: 2, 2+8, 2+16, 2+24, 2+32 ...
        if i == 3:
            current_start_col = initial_start_col + i * num_cols_merge
            current_end_col = current_start_col + num_cols_merge_6 - 1
        else:    
            current_start_col = initial_start_col + i * num_cols_merge
            current_end_col = current_start_col + num_cols_merge - 1

        # Получаем ссылку на верхнюю левую ячейку объединенного диапазона
        merged_cell = ws.cell(row=base_start_row, column=current_start_col)

        # Устанавливаем значение римской цифры
        merged_cell.value = roman_value

        # Применяем стили (шрифт, выравнивание, граница)
        merged_cell.font = font_style
        merged_cell.alignment = alignment_style
        merged_cell.border = border_style
        
        # Объединяем ячейки
        ws.merge_cells(start_row=base_start_row, start_column=current_start_col,
                       end_row=end_row_merge, end_column=current_end_col)
    
    
    # --- Двойные границы для строк, справа от таблицы. Определяем константы, которые не меняются в цикле
    base_start_col = cols + 2  # Начинаем с 32 столбца (cols + 2)
    num_rows_merge = 8  # +8 для орбит.группировки из 8 КА
    num_rows_merge_6 = 6  # +6 для орбит.группировки из 6 КА
    num_cols_merge = 2   
    end_col_merge = base_start_col + num_cols_merge - 1
    
    # Устанавливаем ширину столбцов для объединенных ячеек    
    for k in range(base_start_col, base_start_col + 4):
        ws.column_dimensions[get_column_letter(k)].width = 5 # Ширина base_start_col и + 3 столбца - 5 символов
    
    # Начальный столбец для первой объединенной ячейки
    initial_start_row = 2

    # Проходимся по списку римских цифр
    for i, roman_value in enumerate(rim_num):
        # Рассчитываем начальную строку для текущей итерации
        # start_row: 2, 2+8, 2+16, 2+24, 2+32 ...
        if i == 3:
            current_start_row = initial_start_row + i * num_rows_merge
            current_end_row = current_start_row + num_rows_merge_6 - 1
        else:    
            current_start_row = initial_start_row + i * num_rows_merge
            current_end_row = current_start_row + num_rows_merge - 1

        # Получаем ссылку на верхнюю левую ячейку объединенного диапазона
        merged_cell = ws.cell(row=current_start_row, column=base_start_col)

        # Устанавливаем значение римской цифры
        merged_cell.value = roman_value

        # Применяем стили (шрифт, выравнивание, граница)
        merged_cell.font = font_style
        merged_cell.alignment = alignment_style
        merged_cell.border = border_style
        
        # Объединяем ячейки
        ws.merge_cells(start_row=current_start_row, start_column=base_start_col,
                       end_row=current_end_row, end_column=end_col_merge)
                       
    # Сохраняем книгу
    wb.save(excel_filepath)
    print(f"Excel файл '{excel_filepath}' успешно создан.")
    
    
def map_values_to_colors(values_list, colors_list):
    """
    Создает словарь, где ключи - значения из values_list,
    а значения - соответствующие цвета из colors_list.

    Args:
        values_list (list): Список числовых значений.
        colors_list (list): Список ARGB строк цветов.

    Returns:
        dict: Словарь, отображающий значения на цвета.
              Если длины списков не совпадают, zip() остановится по кратчайшему.
              Если в values_list есть повторяющиеся значения, в словаре будет
              сохранено последнее соответствующее значение цвета.
    """
    if len(values_list) != len(colors_list):
        print("Предупреждение: Длины списков 'values_list' и 'colors_list' не совпадают. "
              "Создание словаря будет остановлено по длине кратчайшего списка.")

    value_to_color_map = {}

    # zip() объединяет элементы из двух списков в кортежи: (value, color)
    for value, color in zip(values_list, colors_list):
        value_to_color_map[value] = color

    return value_to_color_map
    
    
# --- Главная часть скрипта ---

if __name__ == "__main__":
    # генерируем список чисел от 0 до 100
    values_list = list()
    for i in range(101):
        values_list.append(i)
    # Генерируем цвета от 0 до 100
    colors_list = generate_red_to_green_gradient_argb()
    
    # Создаем словарь (глобальная переменная)
    mapped_data = map_values_to_colors(values_list, colors_list)
        
    # моя функция для генерации файла с процентами видимости КА. 
    generate_visability_data_file(DATA_FILE, NUM_ROWS, NUM_COLS, NUM_FILES)
    #else:
 
       #  Загружаем данные из файла
    table_data = load_data_from_file(DATA_FILE, NUM_ROWS, NUM_COLS)

    if table_data is not None:
        # Создаем Excel таблицу
        create_excel_table(table_data, EXCEL_FILE, NUM_ROWS, NUM_COLS)
    else:
        print("Не удалось загрузить данные. Excel файл не будет создан.")